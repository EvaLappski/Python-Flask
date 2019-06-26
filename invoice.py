import openpyxl
import sys
from flask import Flask
app = Flask(__name__)
import json
from flask import request
from flask_cors import CORS
CORS(app)

book = openpyxl.load_workbook('DatabaseEva.xlsx')
sheet = book.active
sheets = book.sheetnames
file = open('invoice.txt', 'a')


customer_list = []
invoice_list = []
line_list = []
product_list = []
values=[]


class Build ():
	@staticmethod
	def check_for_blanks():
		for sheet in book:
			for row in sheet.iter_rows():
				empty_cell_count = 0
				for cell in row:
					if cell.value is None:
						empty_cell_count += 1
						print("There is a blank cell in:", sheet, cell)
			print(f"Finished checking for blanks in {sheet}")

	@staticmethod
	def build_customer_list():
		frequency={}
		for row in book[sheets[0]].iter_rows():
			for cell in row:
				ref = book[sheets[0]].cell(row=cell.row, column=2).value
				if ref not in values:
					values.append(ref)	
		# print('Unique Values', values)	
		for row in book[sheets[0]].iter_rows():
			for cell in row:
				obj={}
				CustomerNo = book[sheets[0]].cell(row=cell.row, column=1).value
				Customer = book[sheets[0]].cell(row=cell.row, column=2).value
				Address = book[sheets[0]].cell(row=cell.row, column=3).value
				Phone = book[sheets[0]].cell(row=cell.row, column=4).value
				Contact = book[sheets[0]].cell(row=cell.row, column=5).value
				obj['CustomerNo'] = CustomerNo
				obj['Customer'] = Customer
				obj['Address'] = Address
				obj['Phone'] = Phone
				obj['Contact'] = Contact
			frequency[Customer] = frequency.get(Customer, 0) +1
			frequency[CustomerNo] = frequency.get(CustomerNo, 0) +1
			if frequency[Customer] > 1: 
				print ("Duplicate Customer Name:", Customer)	
			if frequency[CustomerNo] > 1:
				print("Duplicate Customer Number:", CustomerNo)
			else:
				customer_list.append(obj)	
		# print("TESTING", customer_list)		
		return customer_list

	@staticmethod
	def build_invoice_list():
		frequency={}
		for row in book[sheets[1]].iter_rows():
			for cell in row:
				obj={}
				InvoiceNo = book[sheets[1]].cell(row=cell.row, column=1).value
				obj['InvoiceNo'] = InvoiceNo
				obj['CustomerNo'] = book[sheets[1]].cell(row=cell.row, column=2).value
				obj['Date'] = book[sheets[1]].cell(row=cell.row, column=3).value
			frequency[InvoiceNo] = frequency.get(InvoiceNo, 0) + 1
			
			if frequency[InvoiceNo] > 1:
				print ("Duplicate Invoice #:", InvoiceNo)	
			else:
				invoice_list.append(obj)	
		# print("test", invoice_list[1])
		return invoice_list		
	
	@staticmethod
	def build_line_list():
		frequency={}
		for row in book[sheets[2]].iter_rows():
			for cell in row:
				obj={}
				obj['InvoiceNo'] = book[sheets[2]].cell(row=cell.row, column=1).value
				obj['ProductNo'] = book[sheets[2]].cell(row=cell.row, column=2).value
				obj['Units'] = book[sheets[2]].cell(row=cell.row, column=3).value
			line_list.append(obj)
		# print(line_list)
		# print(line_list[1])	
		return line_list	

	@staticmethod
	def build_product_list():
		frequency={}
		for row in book[sheets[3]].iter_rows():
			for cell in row:
				obj={}
				ProductNo =  book[sheets[3]].cell(row=cell.row, column=1).value
				Product = book[sheets[3]].cell(row=cell.row, column=2).value
				UnitPrice = book[sheets[3]].cell(row=cell.row, column=3).value
				obj['ProductNo'] = ProductNo
				obj['Product'] = Product
				obj['UnitPrice'] = UnitPrice
			frequency[Product] = frequency.get(Product, 0) +1
			frequency[ProductNo] = frequency.get(ProductNo, 0) +1
			if frequency[Product] > 1  :
				print ("Duplicate Product Name:", Product)		
			if 	frequency[ProductNo] > 1 :
				print ("Duplicate Product Number:", ProductNo)				
			else:	
				product_list.append(obj)
		# print(product_list[1])
		return product_list
	@staticmethod
	def build_the_master_lists():
		Build.build_customer_list()
		Build.build_invoice_list()
		Build.build_line_list()
		Build.build_product_list()

Build()
Build.check_for_blanks()
Build.build_the_master_lists()	


sys.stdout = open('invoice.txt','wt')
class Create ():
	@staticmethod
	def create_invoice(inv):
		for obj in invoice_list:
			if obj['InvoiceNo'] == int(inv):
				cust_ref = obj['CustomerNo']
				print('_________________________')
				print('Date:', obj["Date"])
				print('Invoice #:', inv)
				print('Customer #:', cust_ref)
				print('_________________________')				
		for obj in customer_list:
			if obj['CustomerNo']== cust_ref:
				print('Customer:',obj['Customer'])
				print('Address:', obj['Address'])
				print('Phone:', obj['Phone'])
				print('Contact:', obj['Contact'])
				print('_________________________')		
		grand_total = 0

		for obj in line_list:
			if obj['InvoiceNo'] == int(inv):
				product_ref = obj['ProductNo']
				product_purchase = obj['Units']
				for obj in product_list:
					if obj['ProductNo'] == product_ref:
						total = '{:.2f}'.format(product_purchase * obj['UnitPrice'])
						grand_total += float(total)
						print('__________________________________________________________')
						print(obj["Product"],'------ x', product_purchase, '------ $','{:.2f}'.format(obj['UnitPrice']), '------ ', 'TOTAL: $', total)		
						print('__________________________________________________________')
		print('Grand Total: $', grand_total)	

Create.create_invoice(5)


@app.route("/info/")
def info():
	return json.dumps(customer_list)


if __name__ == '__main__':
	app.run(host='0.0.0.0', debug=True)
	