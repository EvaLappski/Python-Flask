from openpyxl import Workbook

book = Workbook()
ws = book.active
ws1 = book.create_sheet('Customers')
ws1.title = "Customers"
ws2 = book.create_sheet('Invoice Number')
ws2.title = "Invoice Number"
ws3 = book.create_sheet('Products')
ws3.title = "Products"
ws4 = book.create_sheet('Line Items')
ws4.title = "Line Items"
print(book.sheetnames)

# CUSTOMER WORKSHEET

ws1['A1'] = 'Customer Number'
ws1['B1'] = 'Customer'
ws1['C1'] = 'Address'
ws1['D1'] = 'Phone Number'
ws1['E1'] = 'Contact Name'
rows = [
    (1223,'Customer X','Edmonton, AB', '587-876-9000', 'Shane McBane'),
    (11234,'HC PIPER', 'Calgary, AB', '403-890-7754', 'Dima Parker'),
    (3467,'BMP MECHANICAL','Calgary, AB','403-619-9773','Jeffrey Lapp'),
    (1223,'Customer Y','Edmonton, AB', '587-876-9000', 'Shane McBane'),
    (1227,'Customer X','Edmonton, AB', '587-876-9000', 'Shane McBane')
]
for row in rows:
    ws1.append(row)
for row in ws1.iter_rows(values_only=True):
			print(row)

# INVOICE WORKSHEET

ws2['A1'] = 'Invoice Number'
ws2['B1'] = 'Customer Number'
ws2['C1'] = 'Date'
rows = [
    (1, 1223,'01/01/2018'),
    (2, 11234, '02/02/2019'),
    (3, 3467,'03/03/2019'),
    (4, 1223,'01/09/2018'),
    (4, 3467,'01/09/2018'),
]
for row in rows:
    ws2.append(row)
for row in ws2.iter_rows(values_only=True):
			print(row)

# LINE ITEMS WORKSHEET
ws3['A1'] = 'Invoice Number'
ws3['B1'] = 'Product'
ws3['C1'] = 'Units'
rows = [
    (2, 'Wash',1),
    (2, 'Wash', 1),
    (2, 'Rinse', 1),
    (2, 'Repeat', 1),
    (4, 'Total Package',1),
]

for row in rows:
    ws3.append(row)
for row in ws3.iter_rows(values_only=True):
			print(row)

# PRODUCT LIST WORKSHEET
ws4['A1'] = 'Product Number'
ws4['B1'] = 'Product Name'
ws4['C1'] = 'Unit Cost'
rows = [
    (1, 'Wash', 50),
    (2, 'Rinse', 30),
    (3, 'Repeat', 10),
    (4,'Total Package', 100),
]

for row in rows:
    ws4.append(row)
for row in ws4.iter_rows(values_only=True):
			print(row)


book.save('InvoiceDummy.xlsx')

# CREATE A NEW BLANK FILE TO MERGE INTO
# master_file = "merge_file.xlsx"

# book = openpyxl.Workbook()
# wscustomer = book.active
# wscustomer.title = 'Customers'
# wsinv = book.create_sheet('Invoice')
# wsline = book.create_sheet('Line Items')
# wsproduct = book.create_sheet('Products')
# print(book, book.sheetnames)

# READ ONE EXCEL WORKBOOK AT A TIME AND MERGE DATA WITHIN TO THE BLANK FILE
# book = openpyxl.load_workbook('DatabaseV3test.xlsx')
# book2 = openpyxl.load_workbook('DatabaseDima.xlsx')
# sheet = book.active
# sheets = book.sheetnames

# wscustomer['A1'] = "Customer Number!!!"
# wscustomer['B1'] = "Customer Name!!!"
# wscustomer['C1'] = "Address!!"
# wscustomer['D1'] = "Phone!!"
# wscustomer['E1'] = "Contact!!"


# def merge_database(book,master):
# 	# load_workbook(file)
# 	for row in book[sheets[0]].iter_rows():	
# 		for cell in row:
# 			CustomerNo = book[sheets[0]].cell(row=cell.row, column=1).value
# 			# print(cell.row)
# 			Customer = book[sheets[0]].cell(row=cell.row, column=2).value
# 			Address = book[sheets[0]].cell(row=cell.row, column=3).value
# 			Phone = book[sheets[0]].cell(row=cell.row, column=4).value
# 			Contact = book[sheets[0]].cell(row=cell.row, column=5).value
# 		print('test',CustomerNo, Customer,Address, Phone, Contact)
# 		rows = (CustomerNo, str(Customer),str(Address), str(Phone), str(Contact))
# 		print(rows)	# if cell.row == 1:
# 		# for row in rows:	# else:
		# 	wscustomer.append(row)

# def merge_another_database(file, master):	
# 	for row in book2[sheets[0]].iter_rows():	
# 		for cell in row:
# 			CustomerNo = book2[sheets[0]].cell(row=cell.row, column=1).value
# 			Customer = book2[sheets[0]].cell(row=cell.row, column=2).value
# 			Address = book2[sheets[0]].cell(row=cell.row, column=3).value
# 			Phone = book2[sheets[0]].cell(row=cell.row, column=4).value
# 			Contact = book2[sheets[0]].cell(row=cell.row, column=5).value
# 		if cell.row == 1:
# 			pass
# 		else:
# 			print(CustomerNo, Customer, Address, Phone, Contact)	
# merge_database(book,'master_file.xlsx')
# # merge_another_database(book2, 'master_file.xlsx')
# book.save("master_file.xlsx")